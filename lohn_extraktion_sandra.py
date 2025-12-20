"""
Lohn-Extraktion fuer Sandra Stolarczyk
Extrahiert Brutto/Netto-Daten und Unterbrechungen aus DATEV-Lohnabrechnungen
"""

import fitz  # PyMuPDF
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

# Konfiguration
PDF_BASE_PATH = r"Z:\Intern\Steuerberater\GmbH\Lohnauswertung"
YEARS = ["2024", "2025"]
SEARCH_NAME = "Sandra"
OUTPUT_FILE = r"Z:\Intern\Steuerberater\GmbH\Lohnauswertung\Sandra_Lohnuebersicht.xlsx"


def is_lohnabrechnung_page(text):
    """Prueft ob die Seite eine Brutto/Netto-Abrechnung ist"""
    has_abrechnung = "Abrechnung der Brutto/Netto" in text or "Brutto/Netto-Bez" in text
    has_sandra = SEARCH_NAME in text
    has_gehalt = "Festbezug" in text or "Lohn/Gehalt" in text
    has_netto = "Netto-Verdienst" in text
    return has_abrechnung and has_sandra and has_gehalt and has_netto


def extract_month_year(text):
    """Extrahiert Monat und Jahr aus dem Abrechnungstext"""
    match = re.search(r'f.r\s+(\w+)\s+(\d{4})', text)
    if match:
        return match.group(1), match.group(2)
    return None, None


def extract_unterbrechungen(text):
    """Extrahiert Unterbrechungszeitraeume"""
    # Suche nach Unterbrechung mit mehrzeiligem Match
    match = re.search(r'Unterbrechung:\s*([0-9]{2}\.-[0-9]{2}\.[0-9]{2}\.[0-9]{2}(?:\s*und\s*\n?[0-9]{2}\.-[0-9]{2}\.[0-9]{2}\.[0-9]{2})?)', text)
    if match:
        return match.group(1).replace('\n', ' ').strip()
    return ""


def parse_currency(value_str):
    """Konvertiert deutschen Waehrungsstring zu Float"""
    if not value_str:
        return None
    try:
        clean = value_str.replace('.', '').replace(',', '.')
        return float(clean)
    except:
        return None


def extract_data_from_page(text):
    """Extrahiert alle relevanten Daten aus einer Seite"""
    data = {}

    # Monat und Jahr
    monat, jahr = extract_month_year(text)
    data['Monat'] = monat
    data['Jahr'] = jahr

    # Unterbrechungen
    data['Unterbrechung'] = extract_unterbrechungen(text)

    # Festbezug Lohn/Gehalt - klar definiertes Muster
    festbezug_match = re.search(r'Festbezug Lohn/Gehalt\s*\n\s*L\s*L\s*J\s*\n\s*(\d{1,3}(?:\.\d{3})*,\d{2})', text)
    if festbezug_match:
        data['Festbezug'] = parse_currency(festbezug_match.group(1))
    else:
        data['Festbezug'] = None

    # Finde alle Betraege im Format X.XXX,XX oder XXX,XX
    all_amounts = re.findall(r'(\d{1,3}(?:\.\d{3})*,\d{2})', text)
    amounts_float = [parse_currency(a) for a in all_amounts if parse_currency(a) is not None]

    # Gesamt-Brutto: Suche nach dem Muster - erscheint nach den Brutto-Bezuegen
    # Der Wert ist typischerweise zwischen 1000 und 5000, erscheint nach 50,00 (Gutschein)
    brutto_idx = text.find('Gesamt-Brutto')
    if brutto_idx > 0:
        # Suche den Betrag der zwischen Brutto-Positionen und Steuerabzuegen steht
        # Typisch: nach "50,00" kommt "2.954,00" als Gesamt-Brutto
        gutschein_match = re.search(r'Gutschein\s*\n\s*F\s*F\s*J\s*\n\s*(\d{1,3}(?:\.\d{3})*,\d{2})\s*\n\s*(\d{1,3}(?:\.\d{3})*,\d{2})', text)
        if gutschein_match:
            data['Gesamt_Brutto'] = parse_currency(gutschein_match.group(2))
        else:
            # Fallback: suche groessten Betrag zwischen 1000 und 5000
            brutto_candidates = [a for a in amounts_float if 1000 < a < 5000]
            if brutto_candidates:
                data['Gesamt_Brutto'] = max(brutto_candidates)
            else:
                data['Gesamt_Brutto'] = None
    else:
        data['Gesamt_Brutto'] = None

    # Netto-Verdienst und Auszahlungsbetrag - stehen am Ende des Textes
    # Muster: ...311,93\n1.701,77\n... und ganz am Ende 1.651,77
    end_amounts = re.findall(r'(\d{1,3}(?:\.\d{3})*,\d{2})', text[-500:])
    end_floats = [parse_currency(a) for a in end_amounts if parse_currency(a)]

    # Die letzten beiden grossen Betraege (>500) sind Netto und Auszahlung
    large_end = [a for a in end_floats if a > 500]
    if len(large_end) >= 2:
        # Der vorletzte ist Netto, der letzte ist Auszahlung
        data['Netto_Verdienst'] = large_end[-2]
        data['Auszahlung'] = large_end[-1]
    elif len(large_end) == 1:
        data['Netto_Verdienst'] = large_end[0]
        data['Auszahlung'] = large_end[0]
    else:
        data['Netto_Verdienst'] = None
        data['Auszahlung'] = None

    return data


def count_unterbrechung_tage(unterbrech_str):
    """Zaehlt die Anzahl der Unterbrechungstage"""
    if not unterbrech_str:
        return 0

    tage = 0
    # Muster: XX.-XX.MM.YY (Zeitraum)
    ranges = re.findall(r'(\d{2})\.-(\d{2})\.', unterbrech_str)
    for start, end in ranges:
        tage += int(end) - int(start) + 1

    return tage


def process_pdfs():
    """Verarbeitet alle PDFs und extrahiert Sandra-Lohnabrechnungen"""
    all_data = []

    for year in YEARS:
        year_path = Path(PDF_BASE_PATH) / year
        if not year_path.exists():
            print(f"Ordner nicht gefunden: {year_path}")
            continue

        pdf_files = sorted(year_path.glob("Auswertungen*.pdf"))
        print(f"\n=== {year}: {len(pdf_files)} PDFs gefunden ===")

        for pdf_file in pdf_files:
            print(f"  Verarbeite: {pdf_file.name}")
            try:
                doc = fitz.open(str(pdf_file))
                for page_num in range(len(doc)):
                    page = doc[page_num]
                    text = page.get_text()

                    if is_lohnabrechnung_page(text):
                        data = extract_data_from_page(text)
                        if data['Monat'] and data['Jahr']:
                            print(f"    -> {data['Monat']} {data['Jahr']}: Festbez={data.get('Festbezug')}, Brutto={data.get('Gesamt_Brutto')}, Netto={data.get('Netto_Verdienst')}")
                            data['Quelldatei'] = pdf_file.name
                            data['Seite'] = page_num + 1
                            all_data.append(data)

                doc.close()

            except Exception as e:
                print(f"    FEHLER: {e}")

    return all_data


def create_excel(data):
    """Erstellt Excel-Datei mit den extrahierten Daten"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sandra Lohnuebersicht"

    headers = ['Jahr', 'Monat', 'Unterbrechung', 'Tage', 'Festbezug', 'Gesamt-Brutto',
               'Netto-Verdienst', 'Auszahlung', 'Quelldatei']

    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border

    monate_order = {'Januar': 1, 'Februar': 2, 'Feb': 2, 'Maerz': 3, 'M\u00e4rz': 3, 'April': 4,
                    'Mai': 5, 'Juni': 6, 'Juli': 7, 'August': 8, 'September': 9,
                    'Oktober': 10, 'November': 11, 'Dezember': 12}

    sorted_data = sorted(data, key=lambda x: (
        x.get('Jahr', '0'),
        monate_order.get(x.get('Monat', ''), 0)
    ))

    for row_num, record in enumerate(sorted_data, 2):
        unterbrech = record.get('Unterbrechung', '')
        tage = count_unterbrechung_tage(unterbrech)

        ws.cell(row=row_num, column=1, value=record.get('Jahr'))
        ws.cell(row=row_num, column=2, value=record.get('Monat'))
        ws.cell(row=row_num, column=3, value=unterbrech)
        ws.cell(row=row_num, column=4, value=tage if tage > 0 else None)
        ws.cell(row=row_num, column=5, value=record.get('Festbezug'))
        ws.cell(row=row_num, column=6, value=record.get('Gesamt_Brutto'))
        ws.cell(row=row_num, column=7, value=record.get('Netto_Verdienst'))
        ws.cell(row=row_num, column=8, value=record.get('Auszahlung'))
        ws.cell(row=row_num, column=9, value=record.get('Quelldatei'))

        for col in range(1, 10):
            ws.cell(row=row_num, column=col).border = thin_border

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 25

    for row in range(2, len(sorted_data) + 2):
        for col in [5, 6, 7, 8]:
            ws.cell(row=row, column=col).number_format = '#,##0.00 €'

    wb.save(OUTPUT_FILE)
    print(f"\nExcel gespeichert: {OUTPUT_FILE}")

    return sorted_data


def analyze_deductions(data):
    """Analysiert die Abzuege bei Unterbrechungen"""
    print("\n" + "=" * 100)
    print("ANALYSE: Abzuege bei Unterbrechung (Brutto-Abzug pro Tag)")
    print("=" * 100)

    # Finde vollen Festbezug (Monat ohne Unterbrechung)
    full_months = [r for r in data if not r.get('Unterbrechung') and r.get('Festbezug')]
    if not full_months:
        print("Kein Monat ohne Unterbrechung gefunden!")
        return

    voller_festbezug = full_months[0]['Festbezug']
    print(f"\nErwarteter Tagessatz laut Beispiel: 104,00 EUR Brutto/Tag")
    print(f"Voller Festbezug (ohne Unterbrechung): {voller_festbezug:,.2f} EUR\n".replace(',', 'X').replace('.', ',').replace('X', '.'))

    print("{:<15} {:<30} {:>6} {:>14} {:>14} {:>12} {:>10}".format(
        'Monat', 'Unterbrechung', 'Tage', 'Festbezug', 'Abzug Brutto', 'EUR/Tag', 'Status'))
    print("-" * 105)

    for record in data:
        unterbrech = record.get('Unterbrechung', '')
        if unterbrech and record.get('Festbezug'):
            tage = count_unterbrechung_tage(unterbrech)
            festbezug = record['Festbezug']
            abzug = voller_festbezug - festbezug

            if tage > 0:
                pro_tag = abzug / tage
                erwartet = tage * 104.00
                diff = abs(abzug - erwartet)

                monat_str = "{} {}".format(record['Monat'], record['Jahr'])
                unterbrech_short = unterbrech[:27] + "..." if len(unterbrech) > 30 else unterbrech

                if diff < 5:
                    status = "OK"
                else:
                    status = "DIFF {:.0f}".format(abzug - erwartet)

                line = "{:<15} {:<30} {:>6} {:>14,.2f} {:>14,.2f} {:>12,.2f} {:>10}".format(
                    monat_str, unterbrech_short, tage, festbezug, abzug, pro_tag, status)
                print(line.replace(',', 'X').replace('.', ',').replace('X', '.'))


if __name__ == "__main__":
    print("=" * 80)
    print("Lohn-Extraktion fuer Sandra Stolarczyk")
    print("=" * 80)

    data = process_pdfs()

    if data:
        print(f"\n{len(data)} Lohnabrechnungen gefunden")
        sorted_data = create_excel(data)
        analyze_deductions(sorted_data)
    else:
        print("\nKeine Daten gefunden!")
